import csv
import os
import logging
import pandas as pd
import requests
import openpyxl
from openpyxl.styles import Font
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse, HttpResponseBadRequest
from django.views.decorators.csrf import csrf_exempt, csrf_protect
from django.db import connection
from django.core.files.storage import FileSystemStorage
from django.views.decorators.http import require_GET
from .models import Tournament, Player, Set  # Elimina MatchCharacter de la importación

# En algunos despliegues el modelo MatchCharacter no existe. Definimos
# un símbolo por defecto para evitar errores estáticos y resolver en tiempo
# de ejecución dentro de bloques try/except.
MatchCharacter = None
from .forms import UploadFileForm, PlayerForm, TournamentForm, SetForm

from .api.setByTournament import get_sets_by_event
from .api.eventInfo import get_event_info, get_tournaments_by_country
from django.views.decorators.http import require_GET

# Importa las vistas separadas
from .views_players import *
from .views_torneos import *
from .views_sets import *

from .api.location_mapping import get_department_by_city, get_region_by_department
from .api.getTournamentDetails import get_tournament_details

import re

logger = logging.getLogger(__name__)

# Constantes
STARTGG_URL = "https://api.start.gg/gql/alpha"
STARTGG_KEY = "204bdde1bb958e691497fa76febad15d"

if not STARTGG_KEY:
    raise EnvironmentError("La variable de entorno 'REACT_APP_STARTGG_API_KEY' no está configurada.")

# --- Home Views ---
def home_page(request):
    return render(request, 'index.html')

def home(request):
    query = request.GET.get('q')
    players = Player.objects.filter(nickname__icontains=query) if query else []
    return render(request, 'home-consultas.html', {'players': players, 'query': query})

def player_detail(request, gamertag):
    player = get_object_or_404(Player, gamertag=gamertag)
    sets_as_p1 = Set.objects.filter(player_1=player)
    sets_as_p2 = Set.objects.filter(player_2=player)
    all_sets = sets_as_p1.union(sets_as_p2).order_by('-id')
    # Algunos despliegues no tienen el modelo MatchCharacter definido.
    # Intentamos usarlo, y si no está disponible devolvemos lista vacía.
    try:
        characters_used = (
            MatchCharacter.objects
            .filter(player=player)
            .select_related('character')
            .order_by('game_number')
        )
    except Exception:
        characters_used = []
    context = {
        'player': player,
        'sets': all_sets,
        'characters_used': characters_used,
    }
    return render(request, 'player_detail.html', context)

def home_consultas(request):
    return render(request, 'consultas/home-consultas.html')

# --- API y utilidades generales ---
def obtener_event_id(tournament_name, event_name):
    return f"{tournament_name}_{event_name}_id"

def extract_names_from_url(url):
    """
    Extrae tournament_name y event_name de un link de start.gg
    """
    match = re.search(r'/tournament/([^/]+)/event/([^/]+)/?$', url)
    if match:
        return match.group(1), match.group(2)
    return None, None

def get_event_id_page(request):
    if request.method == 'POST':
        url = request.POST.get('event_url', '').strip()
        tournament_name = request.POST.get('tournament_name', '').strip()
        event_name = request.POST.get('event_name', '').strip()

        # Si se proporciona el link, extraer los nombres
        if url:
            tournament_name, event_name = extract_names_from_url(url)

        if not tournament_name or not event_name:
            return JsonResponse({'error': 'Faltan datos para obtener el ID'}, status=400)

        event_id = obtener_event_id(tournament_name, event_name)
        if event_id:
            return JsonResponse({'event_id': event_id})
        else:
            return JsonResponse({'error': 'Event ID no encontrado'}, status=404)
    return render(request, 'consultas/consultas-api/get_event_id_page.html')

def get_result_page(request):
    return render(request, 'consultas/consultas-api/get_result_page.html')

def get_sets_by_tournament(request, limit=5):
    return render(request, 'consultas/consultas-api/get_sets_by_tournament.html')

@csrf_exempt
def get_event_id_view(request):
    if request.method == 'POST':
        tournament_name = request.POST.get('tournament_name')
        event_name = request.POST.get('event_name')

        if not tournament_name or not event_name:
            return JsonResponse({'error': 'Faltan parámetros o son inválidos'}, status=400)

        event_slug = f"tournament/{tournament_name}/event/{event_name}"
        
        headers = {
            'Content-Type': 'application/json',
            'Accept': 'application/json',
            'Authorization': f'Bearer {STARTGG_KEY}'
        }
        
        query = """
        query EventQuery($slug: String) {
            event(slug: $slug) {
                id
                name
            }
        }
        """
        
        variables = {
            'slug': event_slug
        }
        
        try:
            response = requests.post(
                STARTGG_URL,
                headers=headers,
                json={'query': query, 'variables': variables}
            )
            
            response.raise_for_status()
            
            data = response.json()
            print('API response:', data)

            if 'errors' in data:
                print('API returned errors:', data['errors'])
                return JsonResponse({'error': 'Error in API response'}, status=400)

            if 'data' not in data or 'event' not in data['data']:
                print('Response structure:', data)
                return JsonResponse({'error': 'Invalid response structure'}, status=400)

            event_id = data['data']['event']['id']
            return JsonResponse({'event_id': event_id})
        
        except requests.exceptions.RequestException as e:
            print('Error al obtener el ID del evento:', e)
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Invalid method'}, status=405)

@csrf_exempt
def get_sets_by_tournament_view(request):
    if request.method == 'POST':
        event_id = request.POST.get('event_id')
        try:
            sets = get_sets_by_event(event_id)
            return JsonResponse(sets, safe=False)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    return render(request, 'consultas/consultas-api/get_sets_by_tournament.html')

def get_sets_by_tournament(event_id, limit):
    if not event_id or not limit:
        raise ValueError("Both 'event_id' and 'limit' must be provided.")

    sets = []
    page_number = 1
    total_sets = 0

    while total_sets < limit:
        query = """
        query EventSets($eventId: ID!, $page: Int!, $perPage: Int!) {
            event(id: $eventId) {
                sets(page: $page, perPage: $perPage) {
                    nodes {
                        id
                        displayScore
                        phaseGroup {
                            phase {
                                name
                            }
                        }
                        event {
                            name
                            tournament {
                                name
                            }
                        }
                        games {
                            selections {
                                entrantId
                                characterId
                            }
                        }
                    }
                }
            }
        }
        """
        variables = {
            'eventId': event_id,
            'page': page_number,
            'perPage': limit
        }
        headers = {
            'Content-Type': 'application/json',
            'Authorization': f'Bearer {STARTGG_KEY}'
        }
        response = requests.post(STARTGG_URL, json={'query': query, 'variables': variables}, headers=headers)
        data = response.json()

        if 'errors' in data:
            raise Exception(', '.join(error['message'] for error in data['errors']))

        nodes = data['data']['event']['sets']['nodes']
        for node in nodes:
            if total_sets < limit:
                set_data = {
                    'id': node['id'],
                    'displayScore': node['displayScore'],
                    'phase_name': node['phaseGroup']['phase']['name'],
                    'event_name': node['event']['name'],
                    'tournament_name': node['event']['tournament']['name'],
                    'games': node['games']
                }
                sets.append(set_data)
                total_sets += 1

        if len(nodes) < limit:
            break

        page_number += 1

    return sets

@require_GET
def get_event_info(request):
    event_id = request.GET.get('event_id')
    if not event_id:
        return JsonResponse({'error': 'Event ID is required'}, status=400)

    query = """
    query EventInfo($eventId: ID!) {
        event(id: $eventId) {
            id
            name
            tournament {
                name
            }
            standings(query: {perPage: 1}) {
                nodes {
                    placement
                    entrant {
                        name
                    }
                }
            }
        }
    }
    """
    variables = {'eventId': event_id}
    headers = {
        'Content-Type': 'application/json',
        'Authorization': f'Bearer {STARTGG_KEY}'
    }

    response = requests.post(STARTGG_URL, json={'query': query, 'variables': variables}, headers=headers)
    data = response.json()

    if 'errors' in data:
        return JsonResponse({'error': 'Error fetching event info'}, status=500)

    event = data['data']['event']
    winner_full_name = event['standings']['nodes'][0]['entrant']['name'] if event['standings']['nodes'] else 'N/A'
    
    # Extraer solo el nombre del jugador sin el sponsor
    winner_name = winner_full_name.split('|')[-1].strip() if '|' in winner_full_name else winner_full_name

    info = {
        'name': event['name'],
        'tournament_name': event['tournament']['name'],
        'winner': winner_name
    }
    return JsonResponse(info)

def get_event_info_page(request):
    return render(request, 'consultas/consultas-api/get_event_info.html')

@require_GET
def get_event_info_view(request):
    event_id = request.GET.get('event_id')
    if not event_id:
        return JsonResponse({'error': 'No event_id provided'}, status=400)

    # Consulta a la API de Start.gg para obtener detalles del evento y torneo
    query = """
    query EventAndTournamentInfo($eventId: ID!) {
        event(id: $eventId) {
            id
            name
            startAt
            tournament {
                id
                name
                city
                countryCode
                state
                slug
            }
            standings(query: {perPage: 1}) {
                nodes {
                    placement
                    entrant {
                        name
                    }
                }
            }
            numEntrants
        }
    }
    """
    variables = {'eventId': event_id}
    headers = {
        'Content-Type': 'application/json',
        'Authorization': f'Bearer {STARTGG_KEY}'
    }

    try:
        response = requests.post(STARTGG_URL, json={'query': query, 'variables': variables}, headers=headers)
        response.raise_for_status()
        data = response.json()
    except Exception as e:
        return JsonResponse({'error': f'Error fetching event info: {str(e)}'}, status=500)

    if 'errors' in data:
        return JsonResponse({'error': 'Error fetching event info'}, status=500)

    event = data['data']['event']
    if not event:
        return JsonResponse({'error': 'Event not found'}, status=404)

    # Ganador
    winner_full_name = event['standings']['nodes'][0]['entrant']['name'] if event['standings']['nodes'] else 'N/A'
    winner_name = winner_full_name.split('|')[-1].strip() if '|' in winner_full_name else winner_full_name

    # Fecha
    import datetime
    start_date = 'N/A'
    if event.get('startAt'):
        try:
            start_date = datetime.datetime.utcfromtimestamp(event['startAt']).strftime('%Y-%m-%d')
        except Exception:
            start_date = str(event['startAt'])

    # Construir la URL del torneo
    tournament_url = f"https://www.start.gg/{event['tournament']['slug']}" if event['tournament'].get('slug') else ''

    # Mapear zona y departamento usando las funciones de mapeo
    city = event['tournament'].get('city', 'N/A') if event.get('tournament') else 'N/A'
    department = get_department_by_city(city)
    region = get_region_by_department(department)

    # Obtener asistentes del torneo
    tournament_id = event['tournament'].get('id') if event.get('tournament') else None
    attendees = []
    if tournament_id:
        tournament_details = get_tournament_details(tournament_id)
        attendees = tournament_details.get('attendees', [])

    # Responder con todos los campos requeridos por el frontend
    info = {
        'tournament_name': event['tournament']['name'] if event.get('tournament') else 'N/A',
        'winner': winner_name,
        'attendees': event.get('numEntrants', 'N/A'),
        'region': region,
        'country': event['tournament'].get('countryCode', 'N/A') if event.get('tournament') else 'N/A',
        'department': department,
        'city': city,
        'start_date': start_date,
        'id': event.get('id', event_id),
        'tournament_url': tournament_url,
        'attendees_list': attendees  # Nueva clave para la lista de asistentes
    }
    return JsonResponse(info)

@require_GET
def get_tournaments_by_country_view(request):
    country_code = request.GET.get('country_code')
    per_page = int(request.GET.get('per_page', 10))  # Valor por defecto de 10 si no se proporciona

    if not country_code:
        return JsonResponse({'error': 'Country code is required'}, status=400)

    try:
        tournaments = get_tournaments_by_country(country_code, per_page)
        print(f"Torneos en {country_code}: {tournaments}")  # Agrega esta línea para depuración
        return JsonResponse(tournaments, safe=False)
    except Exception as e:
        print(f"Error en get_tournaments_by_country_view: {e}")  # Agrega esta línea para depuración
        return JsonResponse({'error': str(e)}, status=500)

def get_event_results(request):
    return JsonResponse({"message": "Resultados del evento"})

def download_tournament_info_csv(request, event_id):
    try:
        event_info = get_event_info(event_id)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="event_{event_id}_info.csv"'

    writer = csv.writer(response)
    writer.writerow(['Nombre del Torneo', 'Ganador', 'Asistentes', 'Region', 'País', 'Departamento', 'Ciudad', 'Fecha', 'ID del Evento'])
    writer.writerow([
        event_info.get('tournament_name', 'N/A'),
        event_info.get('winner', 'N/A'),
        event_info.get('attendees', 'N/A'),
        event_info.get('region', 'N/A'),
        event_info.get('country', 'N/A'),
        event_info.get('department', 'N/A'),
        event_info.get('city', 'N/A'),
        event_info.get('start_date', 'N/A'),
        event_id
    ])

    return response

def download_tournament_info_xlsx(request, event_id):
    tournament = Tournament.objects.get(id=event_id)
    
    # Crear un libro de trabajo y una hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Event Info"
    
    # Añadir encabezados
    headers = ["Nombre del Torneo", "Ganador", "Asistentes", "Region", "País", "Departamento", "Ciudad", "Fecha", "ID del Evento"]
    ws.append(headers)
    
    # Añadir datos del torneo
    data = [
        tournament.tournament_name,
        tournament.winner,
        tournament.attendees,
        tournament.zona,
        tournament.pais,
        tournament.departamento,  # Cambiar 'region' a 'departamento'
        tournament.ciudad,
        tournament.date,
        tournament.id
    ]
    ws.append(data)
    
    # Crear una respuesta HTTP con el archivo XLSX
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=event_{event_id}.xlsx'
    wb.save(response)
    
    return response

from .api.getPlayerDetails import (
    get_player_details_with_sets, get_player_info_auto, get_user_info_by_slug,
    get_player_details, search_gamertag_by_sponsor, get_player_by_unique_id
)

def get_player_info_page(request):
    if request.method == 'GET':
        # Si no se proporcionan parámetros, renderiza el HTML
        if 'method' not in request.GET or 'input' not in request.GET:
            return render(request, 'consultas/consultas-api/get_player_info.html')

        # Obtén los parámetros de la solicitud
        search_method = request.GET.get('method', 'id')
        player_input = request.GET.get('input', '').strip()

        logger.debug(f"Parámetros recibidos: method={search_method}, input={player_input}")

        if not player_input:
            return JsonResponse({'error': 'No se proporcionó un valor para buscar.'}, status=400)

        try:
            if search_method == 'id':
                # Búsqueda por ID
                details = get_player_details(player_input)
            elif search_method == 'slug':
                # Búsqueda por Id Slug
                user_info = get_user_info_by_slug(player_input)
                if user_info and 'player' in user_info:
                    player_id = user_info['player']['id']
                    details = get_player_details(player_id)
                else:
                    return JsonResponse({'error': 'Jugador no encontrado.'}, status=404)
            elif search_method == 'id_gamerTag':
                # Búsqueda por Id Gamertag
                details = get_player_details(player_input)
            else:
                return JsonResponse({'error': 'Método de búsqueda no válido.'}, status=400)

            if 'error' in details:
                return JsonResponse({'error': details['error']}, status=404)

            # Añadir campos de redes sociales si no existen
            for key in ['twitter', 'discord', 'twitch']:
                if key not in details:
                    details[key] = None

            return JsonResponse(details)
        except Exception as e:
            logger.exception(f"Error inesperado al obtener información del jugador: {e}")
            return JsonResponse({'error': 'Error interno del servidor'}, status=500)

    return JsonResponse({'error': 'Método HTTP no permitido.'}, status=405)

from .api.getPlayerDetails import search_gamertag_by_sponsor, get_user_info_by_slug, get_player_details

def search_gamertag_view(request):
    if request.method == 'GET' and 'tourney_slug' in request.GET and 'sponsor' in request.GET:
        tourney_slug = request.GET.get('tourney_slug')
        sponsor = request.GET.get('sponsor')
        results = search_gamertag_by_sponsor(tourney_slug, sponsor)
        if 'error' in results:
            return JsonResponse({'error': results['error']}, status=404)
        return JsonResponse(results)
    return render(request, 'consultas/consultas-api/search_gamertag.html')

from .api.getPlayerDetails import get_player_by_unique_id

def get_player_by_id_view(request):
    if request.method == 'GET' and 'unique_id' in request.GET:
        unique_id = request.GET.get('unique_id')
        details = get_player_by_unique_id(unique_id)
        if 'error' in details:
            return JsonResponse({'error': details['error']}, status=404)
        return JsonResponse(details)
    return render(request, 'consultas/consultas-api/get_player_by_id.html')

def generate_player_excel(details):
    """
    Genera un archivo Excel con los detalles del jugador, con formato igual al Excel de ejemplo.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Detalles del Jugador"

    # Encabezados en el orden y formato del Excel (sin tilde)
    headers = ["ID", "GamerTag", "Slug", "Prefijo", "Nombre", "Pais", "Departamento", "Region", "Ciudad", "Twitter", "Discord", "Twitch"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)

    # Usar claves en mayúsculas y sin tilde, como las retorna get_player_details
    slug = details.get("Slug", "N/A")
    if isinstance(slug, str) and slug.startswith("user/"):
        slug = slug[5:]

    data = [
        details.get("ID", "N/A"),
        details.get("GamerTag", "N/A"),
        slug,
        details.get("Prefijo", "N/A"),
        details.get("Nombre", "N/A"),
        details.get("Pais", "N/A"),
        details.get("Departamento", "N/A"),
        details.get("Region", "N/A"),
        details.get("Ciudad", "N/A"),
        details.get("Twitter", ""),
        details.get("Discord", ""),
        details.get("Twitch", ""),
    ]
    for col_num, value in enumerate(data, 1):
        ws.cell(row=2, column=col_num, value=value)

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="player_details.xlsx"'
    wb.save(response)
    return response

def download_player_excel(request):
    """
    Vista para descargar un archivo Excel con los detalles del jugador.
    """
    player_id = request.GET.get("player_id")
    # Validar que player_id sea un número válido y no esté vacío ni sea "undefined"
    if not player_id or player_id in ["undefined", "null", "None"]:
        return JsonResponse({"error": "Se requiere un player_id válido"}, status=400)
    try:
        player_id_int = int(player_id)
    except (ValueError, TypeError):
        return JsonResponse({"error": "Se requiere un player_id válido"}, status=400)

    details = get_player_details(player_id_int)
    if not details or "error" in details:
        return JsonResponse({"error": details["error"] if details and "error" in details else "Jugador no encontrado"}, status=404)

    return generate_player_excel(details)

def consultas_home(request):
    return render(request, 'consultas/consultas-api/consultas_home.html')

def autocomplete_players(request):
    query = request.GET.get('q', '')
    results = []
    if query:
        players = Player.objects.filter(gamertag__icontains=query)[:10]
        results = [{'id': p.id, 'gamertag': p.gamertag} for p in players]
    return JsonResponse(results, safe=False)

def get_set_info_view(request):
    return render(request, 'consultas/consultas-api/get_set_info.html')

from django.http import JsonResponse
from .api.get_set_info import get_set_info

def get_set_info_api(request):
    set_id = request.GET.get('set_id')
    if not set_id:
        return JsonResponse({'error': 'No set_id provided'}, status=400)
    try:
        set_info = get_set_info(set_id)
        return JsonResponse(set_info, safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


def api_health(request):
    """Simple health check for API-only microservice mode."""
    return JsonResponse({'status': 'ok', 'service': 'consultas-api'})



