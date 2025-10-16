import requests

import openpyxl
from openpyxl.styles import Font
from django.http import HttpResponse
from Consultas.api.location_mapping import get_department_by_city, get_zone_by_department, get_region_by_department
import json
from django.http import JsonResponse
import traceback
from django.db import transaction
from django.db import connection
from django.db.utils import OperationalError
from django.db.models import Q
import re

STARTGG_URL = "https://api.start.gg/gql/alpha"
STARTGG_KEY = "3e417c2b55e54203247b7501c15e70ca"

if not STARTGG_KEY:
    raise ValueError("La clave de la API de Start.gg no está configurada.")

# modelos relacionados a ignorar (por nombre de clase / object_name) — reutilizado en validaciones
IGNORE_RELATED_MODELS = {'Character', 'character', 'CharacterModel', 'Character_Ssbu'}

# patrones para ignorar tablas relacionadas por nombre de tabla (lower-case substrings)
IGNORE_RELATED_TABLE_PATTERNS = ('character',)

# Función genérica para realizar consultas a la API de Start.gg
def startgg_query(query, variables):
    headers = {
        'Content-Type': 'application/json',
        'Authorization': f'Bearer {STARTGG_KEY}'
    }
    try:
        response = requests.post(
            STARTGG_URL,
            headers=headers,
            json={'query': query, 'variables': variables}
        )
        response.raise_for_status()
        return response.json()
    except requests.exceptions.RequestException as e:
        return {"error": str(e)}

# Función para obtener información de un usuario usando su slug
def get_user_info_by_slug(slug):
    """
    Obtiene información de un usuario usando su slug.
    """
    query = """
    query UserBySlug($slug: String!) {
      user(slug: $slug) {
        id
        player {
          id
          gamerTag
        }
      }
    }
    """
    variables = {"slug": slug}
    result = startgg_query(query, variables)
    print(f"Respuesta de get_user_info_by_slug para {slug}: {result}")  # Registro para depuración
    try:
        return result['data']['user']
    except (KeyError, TypeError):
        return None

# Función para obtener detalles de un jugador usando su player.id, incluyendo el slug del usuario y redes sociales
def get_player_details(player_id):
    """
    Obtiene detalles de un jugador usando su player.id, con formato de columnas igual al Excel.
    """
    query = """
    query PlayerInfo($id: ID!) {
      player(id: $id) {
        id
        gamerTag
        prefix
        user {
          name
          slug
          location {
            country
            state
            city
          }
          authorizations {
            type
            externalUsername
          }
        }
      }
    }
    """
    variables = {'id': player_id}
    result = startgg_query(query, variables)
    print(f"Respuesta de get_player_details para {player_id}: {result}")  # Registro para depuración
    try:
        player_data = result['data']['player']
        if player_data:
            user = player_data['user']
            # Extraer redes sociales
            twitter = None
            discord = None
            twitch = None
            for auth in user.get('authorizations', []):
                if auth['type'] == 'TWITTER':
                    twitter = auth.get('externalUsername')
                elif auth['type'] == 'DISCORD':
                    discord = auth.get('externalUsername')
                elif auth['type'] == 'TWITCH':
                    twitch = auth.get('externalUsername')
            department = get_department_by_city(user['location'].get('city', ''))
            region = get_region_by_department(department)
            return {
                'ID': player_data['id'],
                'GamerTag': player_data['gamerTag'],
                'Slug': user.get('slug', 'N/A'),
                'Prefijo': player_data.get('prefix', 'N/A'),
                'Nombre': user.get('name', 'N/A'),
                'Pais': user['location'].get('country', 'N/A'),
                'Departamento': department,
                'Region': region,
                'Ciudad': user['location'].get('city', 'N/A'),
                'Twitter': twitter,
                'Discord': discord,
                'Twitch': twitch,
            }
    except (KeyError, TypeError):
        return {"error": "Jugador no encontrado o datos no disponibles"}

# Función para obtener detalles de un jugador junto con sus sets
def get_player_details_with_sets(player_id, per_page=5, page=1):
    """
    Obtiene detalles de un jugador junto con sus sets usando su player.id.
    """
    query = """
    query PlayerWithSets($id: $id, $perPage: Int!, $page: Int!) {
        player(id: $id) {
            id
            gamerTag
            prefix
            user {
                name
                location {
                    country
                    state
                    city
                }
            }
            sets(perPage: $perPage, page: $page) {
                nodes {
                    id
                    displayScore
                    event {
                        id
                        name
                        tournament {
                            id
                            name
                        }
                    }
                }
            }
        }
    }
    """
    variables = {
        'id': player_id,
        'perPage': per_page,
        'page': page
    }
    return startgg_query(query, variables)

# Función para buscar jugadores por su sponsor en un torneo
def search_gamertag_by_sponsor(tourney_slug, sponsor):
    """
    Busca jugadores por su sponsor en un torneo específico.
    """
    query = """
    query PrefixSearchAttendees($tourneySlug: String!, $sponsor: String!) {
        tournament(slug: $tourneySlug) {
            id
            name
            participants(query: {
                filter: {
                    search: {
                        fieldsToSearch: ["prefix"],
                        searchString: $sponsor
                    }
                }
            }) {
                nodes {
                    id
                    gamerTag
                }
            }
        }
    }
    """
    variables = {
        'tourneySlug': tourney_slug,
        'sponsor': sponsor
    }
    result = startgg_query(query, variables)
    try:
        tournament = result['data']['tournament']
        if not tournament:
            return {"error": "Torneo no encontrado"}
        participants = tournament['participants']['nodes']
        return {
            'tournament_name': tournament.get('name', 'N/A'),
            'participants': [
                {'id': p['id'], 'gamerTag': p['gamerTag']} for p in participants
            ]
        }
    except (KeyError, TypeError):
        return {"error": "Error en la respuesta de la API"}

# Función para obtener el player.id asociado a un user.id
def get_player_id_from_user(user_id):
    """
    Obtiene el player.id asociado a un user.id en Start.gg.
    """
    query = """
    query getUserPlayers($id: ID!) {
      user(id: $id) {
        players {
          id
          gamerTag
        }
      }
    }
    """
    variables = {"id": user_id}
    result = startgg_query(query, variables)
    try:
        players = result['data']['user']['players']
        if players:
            return players[0]['id']  # Devuelve el primer player.id encontrado
    except (KeyError, TypeError):
        pass
    return None

# Función principal para obtener información del jugador
def get_player_info_auto(id_input, per_page=5, page=1, method="id"):
    """
    Intenta obtener información del jugador usando player.id, user.slug o gamerTag.
    """
    if method == "id":
        # Búsqueda por ID
        player_data = get_player_details_with_sets(id_input, per_page, page)
        if player_data and 'error' not in player_data and player_data.get('data', {}).get('player'):
            return player_data

        # Búsqueda por slug
        user_info = get_user_info_by_slug(id_input)
        if user_info and 'player' in user_info:
            player_id = user_info['player']['id']
            return get_player_details_with_sets(player_id, per_page, page)

    elif method == "gamerTag":
        # Búsqueda por gamerTag
        players = []  # Replace with actual logic to search players by gamerTag
        if players:
            player_id = players[0]['player']['id']
            return get_player_details_with_sets(player_id, per_page, page)

    return {"error": "Jugador no encontrado o datos no disponibles"}

# Función para obtener información de un usuario y su jugador asociado usando su user.id
def get_player_by_unique_id(unique_id):
    """
    Obtiene información de un usuario y su jugador asociado usando su user.id.
    """
    query = """
    query PlayerByUniqueId($id: ID!) {
        user(id: $id) {
            id
            name
            player {
                id
                gamerTag
                prefix
            }
            location {
                country
                state
                city
            }
        }
    }
    """
    variables = {"id": unique_id}
    result = startgg_query(query, variables)
    try:
        user = result['data']['user']
        if not user:
            return {"error": "Usuario no encontrado"}
        details = {
            'id': user.get('id', 'N/A'),
            'name': user.get('name', 'N/A'),
            'gamerTag': user['player'].get('gamerTag', 'N/A') if user.get('player') else 'N/A',
            'prefix': user['player'].get('prefix', 'N/A') if user.get('player') else 'N/A',
            'country': user['location'].get('country', 'N/A') if user.get('location') else 'N/A',
            'state': user['location'].get('state', 'N/A') if user.get('location') else 'N/A',
            'city': user['location'].get('city', 'N/A') if user.get('location') else 'N/A',
        }
        return details
    except (KeyError, TypeError):
        return {"error": "Error en la respuesta de la API"}

def generate_player_excel(details):
    """
    Genera un archivo Excel con los detalles del jugador, con formato igual al Excel de ejemplo.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Detalles del Jugador"

    # Encabezados en el orden y formato del Excel (sin tilde)
    headers = ["ID", "GamerTag", "Slug", "Prefijo", "Nombre", "Pais", "Departamento", "Region", "Ciudad", "Twitter", "Discord", "Twitch"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)

    # Procesar slug para quitar 'user/' si existe
    slug = details.get("Slug", "N/A")
    if isinstance(slug, str) and slug.startswith("user/"):
        slug = slug[5:]

    # Datos del jugador en el orden correcto y sin tildes
    data = [
        details.get("ID", "N/A"),
        details.get("GamerTag", "N/A"),
        slug,
        details.get("Prefijo", "N/A"),
        details.get("Nombre", "N/A"),
        details.get("Pais", "N/A"),
        details.get("Departamento", "N/A"),
        details.get("Region", "N/A"),
        details.get("Ciudad", "N/A"),
        details.get("Twitter", None),
        details.get("Discord", None),
        details.get("Twitch", None),
    ]
    for col_num, value in enumerate(data, 1):
        ws.cell(row=2, column=col_num, value=value)

    # Crear respuesta HTTP con el archivo Excel
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="player_details.xlsx"'
    wb.save(response)
    return response

def get_attendees_by_tournament_id(tournament_id, per_page=100, page=1):
    """
    Obtiene la lista de asistentes (jugadores) a un torneo por su ID.
    """
    query = """
    query TournamentAttendees($tournamentId: ID!, $perPage: Int!, $page: Int!) {
      tournament(id: $tournamentId) {
        id
        name
        participants(query: {perPage: $perPage, page: $page}) {
          pageInfo {
            totalPages
            total
          }
          nodes {
            id
            gamerTag
            player {
              id
            }
          }
        }
      }
    }
    """
    variables = {
        "tournamentId": tournament_id,
        "perPage": per_page,
        "page": page
    }
    result = startgg_query(query, variables)
    try:
        tournament = result['data']['tournament']
        if not tournament:
            return {"error": "Torneo no encontrado"}
        participants = tournament['participants']['nodes']
        # DEBUG: imprime los participantes y sus player_id
        for p in participants:
            print(f"Participante: {p.get('gamerTag')} - participant_id: {p.get('id')} - player_id: {p.get('player', {}).get('id')}")
        return [
            {
                "participant_id": p.get("id"),
                "gamerTag": p.get("gamerTag"),
                "player_id": p.get("player", {}).get("id")
            }
            for p in participants
        ]
    except (KeyError, TypeError):
        return {"error": "Error en la respuesta de la API"}

def get_player_info_view(request):
    """
    API view: acepta player_id, slug o gamerTag (GET/POST/JSON).
    Devuelve JSON con los detalles o {"error": "..."}.
    """
    # extraer params GET / POST / JSON
    player_id = request.GET.get('player_id') or request.POST.get('player_id')
    slug = request.GET.get('slug') or request.POST.get('slug')
    gamertag = request.GET.get('gamerTag') or request.POST.get('gamerTag') or request.GET.get('gamertag') or request.POST.get('gamertag')
    method = request.GET.get('method') or request.POST.get('method')

    if request.content_type and 'application/json' in (request.content_type or ''):
        try:
            body = json.loads(request.body.decode('utf-8') or "{}")
            player_id = player_id or body.get('player_id') or body.get('id')
            slug = slug or body.get('slug')
            gamertag = gamertag or body.get('gamerTag') or body.get('gamertag')
            method = method or body.get('method')
        except Exception:
            pass

    if player_id:
        # preferimos buscar por player.id
        try:
            # devolver información formateada (sin sets)
            details = get_player_details(int(player_id))
            if isinstance(details, dict) and details.get('error'):
                return JsonResponse({"success": False, "error": details.get('error')}, status=404)
            return JsonResponse({"success": True, "details": details})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)}, status=500)

    if slug:
        # slug puede ser 'user/xxx' o solo 'xxx'
        try:
            # intentar buscar user -> player via get_user_info_by_slug
            user_info = get_user_info_by_slug(slug)
            if not user_info:
                return JsonResponse({"success": False, "error": "User not found for slug"}, status=404)
            player = user_info.get('player')
            if not player or not player.get('id'):
                return JsonResponse({"success": False, "error": "No player associated with user"}, status=404)
            player_id = player['id']
            details = get_player_details(player_id)
            return JsonResponse({"success": True, "details": details})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)}, status=500)

    if gamertag:
        # intentar búsqueda por gamertag vía get_player_info_auto (implementación parcial en el módulo)
        try:
            res = get_player_info_auto(gamertag, method="gamerTag")
            if isinstance(res, dict) and res.get('error'):
                return JsonResponse({"success": False, "error": res.get('error')}, status=404)
            return JsonResponse({"success": True, "data": res})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)}, status=500)

    return JsonResponse({"success": False, "error": "Missing parameter: provide player_id or slug or gamerTag"}, status=400)

# intentar importar el modelo Player (variantes)
try:
    from .models import Player as LocalPlayerModel  # si existe en el mismo paquete
except Exception:
    LocalPlayerModel = None
try:
    if LocalPlayerModel is None:
        from Consultas.models import Player as GlobalPlayerModel
        LocalPlayerModel = GlobalPlayerModel
except Exception:
    # no hay modelo importable; la vista funcionará en modo "dry-run"
    LocalPlayerModel = None

def _find_existing_player(model, details):
    """
    Intentar localizar un jugador existente usando campos comunes (startgg id, slug, gamertag).
    Devuelve instancia o None.
    """
    if not model:
        return None
    field_names = [f.name for f in model._meta.get_fields()]
    # Prioritarios: campo startgg player id (ID), slug, gamertag/nombre
    candidates = []
    if 'ID' in details and details.get('ID'):
        candidates.append(('startgg_player_id', details.get('ID')))
        candidates.append(('startgg_id', details.get('ID')))
        candidates.append(('external_id', details.get('ID')))
    if 'Slug' in details and details.get('Slug'):
        slug = details.get('Slug')
        if isinstance(slug, str) and slug.startswith('user/'):
            slug = slug[5:]
        candidates.append(('slug', slug))
        candidates.append(('user_slug', slug))
    if 'GamerTag' in details and details.get('GamerTag'):
        candidates.append(('gamerTag', details.get('GamerTag')))
        candidates.append(('gamertag', details.get('GamerTag')))
        candidates.append(('name', details.get('GamerTag')))

    for field, val in candidates:
        if not val:
            continue
        if field in field_names:
            try:
                qs = {field: val}
                obj = model.objects.filter(**qs).first()
                if obj:
                    return obj
            except Exception:
                # si el filtro falla, continuar
                continue
    return None

def find_existing_player_by_startgg_id(model, player_id):
    """
    Búsqueda robusta del jugador en la BD usando:
      - coincidencias numéricas en campos startgg_player_id/startgg_id/external_id o pk
      - coincidencias case-insensitive en campos slug/gamerTag/name
      - si no encuentra nada, intenta obtener datos desde start.gg (get_player_details)
        y vuelve a buscar por gamerTag/slug.
    Devuelve instancia o None.
    """
    if model is None or not player_id:
        return None

    field_names = [f.name for f in model._meta.get_fields()]
    q = Q()
    # intentar como entero
    pid_int = None
    try:
        pid_int = int(player_id)
    except Exception:
        pid_int = None

    if pid_int is not None:
        for fname in ('startgg_player_id', 'startgg_id', 'external_id'):
            if fname in field_names:
                q |= Q(**{fname: pid_int})
        # verificar pk directo
        try:
            q |= Q(pk=pid_int)
        except Exception:
            pass

    # coincidencias string (iexact) en varios campos
    for fname in ('startgg_player_id', 'startgg_id', 'external_id', 'slug', 'user_slug', 'gamerTag', 'gamertag', 'name'):
        if fname in field_names:
            q |= Q(**{f"{fname}__iexact": str(player_id)})

    if q:
        try:
            obj = model.objects.filter(q).first()
            if obj:
                return obj
        except Exception:
            # si falla el filtro, continuamos a heurísticas
            pass

    # fallback: si tenemos acceso a start.gg, obtener detalles y buscar por gamerTag/slug
    if 'get_player_details' in globals() and get_player_details is not None:
        try:
            details = get_player_details(int(player_id))
            if isinstance(details, dict) and not details.get('error'):
                gt = details.get('GamerTag')
                slug = details.get('Slug')
                q2 = Q()
                if gt:
                    # buscar gamerTag o name case-insensitive
                    if 'gamerTag' in field_names:
                        q2 |= Q(gamerTag__iexact=gt)
                    if 'gamertag' in field_names:
                        q2 |= Q(gamertag__iexact=gt)
                    if 'name' in field_names:
                        q2 |= Q(name__iexact=gt)
                if slug:
                    s = slug[5:] if isinstance(slug, str) and slug.startswith('user/') else slug
                    if 'slug' in field_names:
                        q2 |= Q(slug__iexact=s)
                    if 'user_slug' in field_names:
                        q2 |= Q(user_slug__iexact=s)
                if q2:
                    try:
                        obj = model.objects.filter(q2).first()
                        if obj:
                            return obj
                    except Exception:
                        pass
        except Exception:
            pass

    return None

def _check_related_tables_exist(model):
    """
    Comprueba que las tablas para modelos relacionados (FK, M2M y through) existan en la BD.
    Devuelve lista de nombres de tablas faltantes (vacía si todo OK).

    Nota: ignoramos temporalmente la tabla 'Character' (y variantes) para permitir crear/actualizar
    Player incluso si aún no has añadido el modelo Character y ejecutado migraciones.
    Si vas a persistir personajes, crea el modelo Character en Consultas/models.py y ejecuta
    makemigrations/migrate; luego quita 'Character' de la lista IGNORE_RELATED_MODELS.
    """
    missing = []
    try:
        existing_tables = connection.introspection.table_names()
    except Exception:
        # si no podemos inspeccionar asumimos que no podemos validar
        return missing

    # usar la constante global de modelos a ignorar
    global IGNORE_RELATED_MODELS

    for f in model._meta.get_fields():
        # relaciones directas a otros modelos (FK, OneToOne, M2M)
        if getattr(f, 'is_relation', False) and getattr(f, 'related_model', None):
            related = f.related_model
            # ignorar modelos abstractos/proxy
            if getattr(related._meta, 'abstract', False) or getattr(related._meta, 'proxy', False):
                continue
            # ignorar si el nombre del modelo está en la lista de excepciones
            related_name = (getattr(related._meta, 'object_name', None) or getattr(related, '__name__', None) or "").strip()
            if related_name in IGNORE_RELATED_MODELS:
                continue
            # ignorar si la tabla relacionada contiene un patrón (por ejemplo 'character') — robusto ante db_table distinto
            try:
                rel_table_candidate = (related._meta.db_table or "").lower()
            except Exception:
                rel_table_candidate = ""
            if any(pat in rel_table_candidate for pat in IGNORE_RELATED_TABLE_PATTERNS):
                continue
            try:
                rt = related._meta.db_table
                if rt not in existing_tables:
                    missing.append(rt)
            except Exception:
                continue
        # para M2M comprobar tabla through (si existe)
        if getattr(f, 'many_to_many', False) and getattr(f, 'remote_field', None):
            through = getattr(f.remote_field, 'through', None)
            if through:
                # ignorar through si corresponde a Character-related through
                through_name = getattr(through._meta, 'object_name', None) or getattr(through, '__name__', None)
                if through_name in IGNORE_RELATED_MODELS:
                    continue
                try:
                    through_table = through._meta.db_table
                    # ignorar throughs que contengan patrones (ej: character)
                    if any(pat in (through_table or "").lower() for pat in IGNORE_RELATED_TABLE_PATTERNS):
                        continue
                    if through_table not in existing_tables:
                        missing.append(through_table)
                except Exception:
                    continue
    # dedupe
    return sorted(list(set(missing)))

def _extract_missing_tables_from_msg(msg):
    """
    Extrae nombres de tabla desde mensajes como:
      "no such table: main.Character"  -> ["character"]
      "no such table: Character"       -> ["character"]
    Devuelve lista dedupeada de nombres (normalizados en lower-case, sin esquema).
    Filtra fragmentos irrelevantes como 'main'.
    """
    if not msg:
        return []
    found = []
    try:
        parts = re.findall(r"no such table:\s*([A-Za-z0-9_.`\"]+)", msg, flags=re.IGNORECASE)
        for p in parts:
            p = p.strip().strip('`"')
            # si viene esquema.tabla tomar solo la tabla
            tbl = p.split('.')[-1].strip()
            if not tbl:
                continue
            tbl_lower = tbl.lower()
            # ignorar nombres de esquema o placeholders
            if tbl_lower in ('main', 'sqlite_master'):
                continue
            found.append(tbl_lower)
    except Exception:
        pass
    return sorted(list(set(found)))

def _sql_literal(v):
    """
    Convierte un valor Python a literal SQL seguro (básico):
     - None -> NULL
     - int/float/bool -> literal
     - str -> 'escaped''string'
    Nota: usado sólo en este fallback; es un escape mínimo suficiente para uso interno.
    """
    if v is None:
        return "NULL"
    if isinstance(v, bool):
        return "1" if v else "0"
    if isinstance(v, (int, float)):
        return str(v)
    # escape simples de comillas simples
    s = str(v).replace("'", "''")
    return f"'{s}'"

def _raw_insert_or_update(obj, created, model, details=None):
    """
    Fallback robusto que:
     - inspecciona las columnas reales de la tabla destino (PRAGMA table_info)
     - construye INSERT/UPDATE sólo con las columnas existentes
     - ejecuta SQL con literales ya escapados (cursor.execute(sql) sin params)
    Devuelve la instancia recuperada por PK si es posible, o el objeto pasado.
    """
    table = model._meta.db_table

    # inspeccionar columnas reales de la tabla en la DB
    existing_cols = []
    try:
        with connection.cursor() as cursor:
            cursor.execute(f'PRAGMA table_info("{table}")')
            info = cursor.fetchall()
            # fila PRAGMA: (cid, name, type, notnull, dflt_value, pk)
            existing_cols = [row[1] for row in info]
    except Exception:
        existing_cols = []

    if not existing_cols:
        raise RuntimeError(f"Cannot inspect table columns for '{table}' (table may not exist)")

    # preparar columnas/valores en base a columnas reales
    cols = []
    vals = []
    for f in model._meta.get_fields():
        if not getattr(f, 'concrete', False) or getattr(f, 'auto_created', False):
            continue
        col = getattr(f, 'column', None)
        att = getattr(f, 'attname', None)
        if not col or att is None:
            continue
        if col not in existing_cols:
            # saltar cualquier columna que no exista físicamente (p. ej. FK a Character)
            continue
        try:
            v = getattr(obj, att)
        except Exception:
            v = None
        cols.append(col)
        vals.append(v)

    if not cols:
        raise RuntimeError(f"No concrete columns available in DB table '{table}' to perform insert/update.")

    # ejecutar INSERT o UPDATE
    try:
        with connection.cursor() as cursor:
            if created:
                cols_sql = ", ".join([f'"{c}"' for c in cols])
                values_sql = ", ".join([_sql_literal(v) for v in vals])
                sql = f'INSERT INTO "{table}" ({cols_sql}) VALUES ({values_sql})'
                cursor.execute(sql)
            else:
                pk_col = model._meta.pk.column
                pk_att = model._meta.pk.attname
                pk_val = getattr(obj, pk_att, None)
                if pk_val is None or pk_col not in existing_cols:
                    # si no hay pk/col PK no es viable UPDATE: intentar INSERT
                    cols_sql = ", ".join([f'"{c}"' for c in cols])
                    values_sql = ", ".join([_sql_literal(v) for v in vals])
                    sql = f'INSERT INTO "{table}" ({cols_sql}) VALUES ({values_sql})'
                    cursor.execute(sql)
                else:
                    set_parts = []
                    for c, v in zip(cols, vals):
                        if c == pk_col:
                            continue
                        set_parts.append(f'"{c}" = {_sql_literal(v)}')
                    set_sql = ", ".join(set_parts) if set_parts else ""
                    sql = f'UPDATE "{table}" SET {set_sql} WHERE "{pk_col}" = {_sql_literal(pk_val)}'
                    cursor.execute(sql)

        # intentar recuperar instancia ORM por PK si existe y la tabla contiene la columna pk
        try:
            pk_att = model._meta.pk.attname
            pk_val = getattr(obj, pk_att, None)
            if pk_val is not None and model._meta.pk.column in existing_cols:
                return model.objects.filter(pk=pk_val).first()
        except Exception:
            pass

        # intenta localizar por heurística usando details (si provisto)
        if details:
            try:
                found = _find_existing_player(model, details)
                if found:
                    return found
            except Exception:
                pass

        return obj
    except Exception as e:
        raise RuntimeError(f"Raw DB save failed: {e}") from e

def _update_or_create_player(model, details):
    """
    Crea o actualiza el modelo Player con los datos disponibles en details.
    Retorna (created_bool, instance).
    """
    if not model:
        raise RuntimeError("Player model not available")

    # comprobar tablas relacionadas antes de intentar guardar (evita errores 'no such table' durante save)
    missing_tables = _check_related_tables_exist(model)
    if missing_tables:
        # No abortamos aquí; intentaremos guardar y manejaremos errores concretos (p. ej. "no such table")
        print(f"Warning: missing related tables for Player model: {missing_tables}. Will attempt save and recover if possible. Run migrations to create these tables.")

    # Verificar que la tabla del modelo exista en la DB (evitar errores 'no such table')
    try:
        existing_tables = connection.introspection.table_names()
    except Exception as e:
        # si no podemos inspeccionar, lanzamos RuntimeError para que el llamador lo capture
        raise RuntimeError(f"Cannot introspect DB tables: {e}")
    
    model_table = model._meta.db_table
    if model_table not in existing_tables:
        raise RuntimeError(f"Database table for model '{model.__name__}' not found: expected table '{model_table}'. Run migrations before creating records.")

    existing = _find_existing_player(model, details)
    created = False
    if existing is None:
        # crear nuevo
        try:
            obj = model()
            created = True
        except Exception as e:
            raise

    else:
        obj = existing

    # mapeo simple de campos desde details a nombres de campo comunes en el modelo
    # solo asigna si el campo existe en el modelo
    field_names = [f.name for f in model._meta.get_fields()]
    mapping = {
        'ID': ['startgg_player_id', 'startgg_id', 'external_id'],
        'GamerTag': ['gamerTag', 'gamertag', 'name'],
        'Slug': ['slug', 'user_slug'],
        'Prefijo': ['prefix'],
        'Nombre': ['name', 'full_name'],
        'Pais': ['country'],
        'Departamento': ['department'],
        'Region': ['region'],
        'Ciudad': ['city'],
        'Twitter': ['twitter'],
        'Discord': ['discord'],
        'Twitch': ['twitch']
    }
    for src_key, target_fields in mapping.items():
        val = details.get(src_key)
        if val is None:
            continue
        for tf in target_fields:
            if tf in field_names:
                try:
                    setattr(obj, tf, val)
                except Exception:
                    pass
                break

    # intentar guardar
    try:
        obj.save()
        return created, obj
    except Exception as save_exc:
        msg = str(save_exc)
        missing_from_error = _extract_missing_tables_from_msg(msg)
 
        if missing_from_error:
            try:
                # para cada tabla faltante, limpiar las columnas FK (attname) que apunten a modelos cuya db_table coincide
                for miss in missing_from_error:
                    m = miss.lower()
                    for f in model._meta.get_fields():
                        if not getattr(f, 'is_relation', False) or getattr(f, 'many_to_many', False):
                            continue
                        related = getattr(f, 'related_model', None)
                        if related is None:
                            continue
                        try:
                            related_table = related._meta.db_table or ''
                        except Exception:
                            related_table = ''
                        rt = related_table.lower()
                        # comparación robusta: equal / contains / endswith
                        if (m == rt) or (rt.endswith(m)) or (m in rt) or (rt in m):
                            attname = getattr(f, 'attname', None)
                            if attname:
                                try:
                                    setattr(obj, attname, None)
                                except Exception:
                                    # ignorar fallos individuales y seguir intentando limpiar otros campos
                                    pass
                # reintentar guardar
                obj.save()
                return created, obj
            except Exception as retry_exc:
                # si reintento con ORM falla, intentar fallback crudo (raw SQL) para persistir columnas concretas
                try:
                    saved_obj = _raw_insert_or_update(obj, created, model, details)
                    return (created, saved_obj)
                except Exception as raw_exc:
                    raise RuntimeError(f"Failed saving Player after clearing relations for missing tables {missing_from_error}: {retry_exc} ; raw fallback error: {raw_exc}") from raw_exc
        
        # si no conseguimos recuperar, normalizar OperationalError a RuntimeError para el llamador
        if isinstance(save_exc, OperationalError):
            raise RuntimeError(f"OperationalError saving Player: {save_exc}") from save_exc
        raise

def sync_players_from_tournament_view(request):
    """
    API view que sincroniza participantes de un torneo a la base de datos de Players.
    Parámetros:
      - tournament_id (GET/POST/JSON)  (requerido)
      - dry_run (opcional) si true no escribe en DB (default false)
    Respuesta JSON con resumen {created: n, updated: n, skipped: n, errors: [...]}
    """
    # extraer params
    tournament_id = request.GET.get('tournament_id') or request.POST.get('tournament_id')
    dry_run = request.GET.get('dry_run') or request.POST.get('dry_run')
    if request.content_type and 'application/json' in (request.content_type or ''):
        try:
            body = json.loads(request.body.decode('utf-8') or "{}")
            tournament_id = tournament_id or body.get('tournament_id')
            dry_run = dry_run or body.get('dry_run')
        except Exception:
            pass

    if not tournament_id:
        return JsonResponse({"success": False, "error": "Missing parameter: tournament_id"}, status=400)

    dry_run_bool = str(dry_run).lower() in ['1', 'true', 'yes'] if dry_run is not None else False

    summary = {"created": 0, "updated": 0, "skipped": 0, "errors": []}
    try:
        attendees = get_attendees_by_tournament_id(tournament_id, per_page=100, page=1)
        if isinstance(attendees, dict) and attendees.get('error'):
            return JsonResponse({"success": False, "error": "Upstream error getting attendees", "details": attendees.get('error')}, status=502)
        if not attendees:
            return JsonResponse({"success": False, "error": "No attendees found or empty list"}, status=404)

        # iterar asistentes
        for a in attendees:
            try:
                player_id = a.get('player_id')
                gamerTag = a.get('gamerTag') or a.get('gamer_tag')
                if not player_id and not gamerTag:
                    summary['skipped'] += 1
                    continue
                # si existe player_id, preferimos obtener detalles desde start.gg
                details = None
                if player_id:
                    try:
                        details = get_player_details(player_id)
                        # get_player_details puede devolver {"error": ...}
                        if isinstance(details, dict) and details.get('error'):
                            # fallback: construir minimal desde attendee
                            details = {
                                'ID': player_id,
                                'GamerTag': gamerTag
                            }
                    except Exception as e:
                        details = {'ID': player_id, 'GamerTag': gamerTag}
                else:
                    details = {'GamerTag': gamerTag}

                # si no hay modelo, no escribimos, solo contamos
                if LocalPlayerModel is None or dry_run_bool:
                    summary['created'] += 0
                    summary['skipped'] += 1 if LocalPlayerModel is None else 0
                    # añadir entry en detalles devueltos para inspección
                    continue

                # otherwise create/update in DB
                with transaction.atomic():
                    try:
                        existing = _find_existing_player(LocalPlayerModel, details)
                        if existing:
                            # actualizar campos
                            created_flag = False
                            created, obj = _update_or_create_player(LocalPlayerModel, details)
                            if not created:
                                summary['updated'] += 1
                            else:
                                summary['created'] += 1
                        else:
                            created, obj = _update_or_create_player(LocalPlayerModel, details)
                            summary['created'] += 1 if created else 0
                    except Exception as e:
                        summary['errors'].append({"participant": a, "error": str(e)})
            except Exception as inner_e:
                summary['errors'].append({"participant": a, "error": str(inner_e)})
                continue

    except Exception as e:
        traceback.print_exc()
        return JsonResponse({"success": False, "error": "Unexpected error", "details": str(e)}, status=500)

    return JsonResponse({"success": True, "dry_run": dry_run_bool, "summary": summary}, status=200)

def ensure_player_view(request):
    """
    API view que recibe player_id (GET/POST/JSON).
    - GET (default): comprueba existencia en BD y devuelve {"success": True, "status": "exists"/"missing", ...}
    - Si missing, también intenta obtener detalles desde start.gg y los devuelve en "details" para inspección.
    - Si se pasa create=true (GET o POST) se intentará crear/actualizar en la BD con los datos obtenidos.
    """
    # extraer player_id y flag create
    player_id = request.GET.get('player_id') or request.POST.get('player_id')
    create_flag = request.GET.get('create') or request.POST.get('create')
    # aceptar JSON body también
    if request.content_type and 'application/json' in (request.content_type or ''):
        try:
            body = json.loads(request.body.decode('utf-8') or "{}")
            player_id = player_id or body.get('player_id') or body.get('id')
            if create_flag is None:
                create_flag = body.get('create')
        except Exception:
            pass

    if not player_id:
        return JsonResponse({"success": False, "error": "Missing parameter: player_id"}, status=400)

    # normalize create flag to boolean
    create_bool = False
    if create_flag is not None:
        create_bool = str(create_flag).lower() in ['1', 'true', 'yes']

    try:
        # intentar localizar en DB si hay modelo
        existing = None
        if LocalPlayerModel is not None:
            try:
                # usar búsqueda robusta por start.gg player id / gamerTag / slug
                existing = find_existing_player_by_startgg_id(LocalPlayerModel, player_id)
                # fallback al método anterior si no hubo resultado
                if existing is None:
                    existing = _find_existing_player(LocalPlayerModel, {'ID': str(player_id)})
            except Exception:
                existing = None

        if existing:
            # ya existe -> devolver info básica
            resp = {"success": True, "status": "exists", "id": existing.pk}
            try:
                resp_fields = {}
                for f in ['gamerTag','name','slug','startgg_player_id']:
                    if hasattr(existing, f):
                        resp_fields[f] = getattr(existing, f)
                if resp_fields:
                    resp['details'] = resp_fields
            except Exception:
                pass
            return JsonResponse(resp, status=200)

        # no existe en BD
        # intentar obtener detalles desde start.gg (no crea todavía)
        details = None
        try:
            details = get_player_details(int(player_id))
        except Exception as e:
            details = {"error": f"Failed to fetch from upstream: {str(e)}"}

        if not create_bool:
            # modo CHECK: devolver missing y los detalles obtenidos (si hay)
            return JsonResponse({"success": True, "status": "missing", "details": details}, status=200)

        # modo CREATE: si no hay modelo para guardar, devolver error
        if LocalPlayerModel is None:
            return JsonResponse({"success": False, "error": "No Player model available to create records", "details": details}, status=500)

        # crear/actualizar en BD usando los detalles obtenidos
        try:
            created, obj = _update_or_create_player(LocalPlayerModel, details)
            status = "created" if created else "updated"
            return JsonResponse({"success": True, "status": status, "id": obj.pk, "details": details}, status=200)
        except RuntimeError as re:
            # errores de esquema (p. ej. tablas relacionadas faltantes) -> devolver mensaje claro
            traceback.print_exc()
            return JsonResponse({"success": False, "error": "DB schema error", "details": str(re)}, status=500)
        except Exception as e:
            traceback.print_exc()
            return JsonResponse({"success": False, "error": "DB error", "details": str(e)}, status=500)

    except Exception as e:
        traceback.print_exc()
        return JsonResponse({"success": False, "error": "Unexpected error", "details": str(e)}, status=500)